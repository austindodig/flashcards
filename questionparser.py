from openpyxl import load_workbook

def merge_cells_with_data(file_path, sheet_name):
    # Load the workbook and sheet
    wb = load_workbook(file_path)
    ws = wb[sheet_name]

    # Get a list of merged ranges
    merged_ranges = list(ws.merged_cells.ranges)

    for merged_range in merged_ranges:
        # Extract bounds of the merged range
        min_col, min_row, max_col, max_row = merged_range.bounds

        # Check if the merged range intersects with one of the target columns
        if min_col == max_col:  # Single-column merge
            column_letter = ws.cell(row=min_row, column=min_col).column_letter

            if column_letter in ['A', 'B', 'C']:  # Targeting Category, Question, Answer
                for col_letter in ['A', 'B', 'C']:
                    if col_letter != column_letter:
                        # Get the range of cells in the column
                        cell_range = ws[f"{col_letter}{min_row}:{col_letter}{max_row}"]

                        # Collect data from all cells in the range
                        cell_data = [str(cell.value) if cell.value is not None else "" for row in cell_range for cell in row]

                        # Concatenate data, separated by spaces
                        concatenated_data = " ".join(cell_data).strip()

                        # Write the concatenated data into the first cell
                        first_cell = ws[f"{col_letter}{min_row}"]
                        first_cell.value = concatenated_data

                        # Merge the cells in the range
                        ws.merge_cells(f"{col_letter}{min_row}:{col_letter}{max_row}")

    # Save the cleaned workbook
    cleaned_file_path = file_path.replace(".xlsx", "_cleaned.xlsx")
    wb.save(cleaned_file_path)
    print(f"Cleaned file saved as {cleaned_file_path}")

# Specify file path and sheet name
file_path = "DPAQuestions.xlsx"  # Replace with your file path
sheet_name = "Sheet1"  # Replace with your sheet name

# Uncomment below to run the function
merge_cells_with_data(file_path, sheet_name)
